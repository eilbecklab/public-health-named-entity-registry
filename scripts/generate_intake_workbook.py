#!/usr/bin/env python3
"""Create a blank intake template from the current versioned workbook."""

from __future__ import annotations

import argparse
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "intake" / "US-FED-DHHS.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "templates" / "phner-intake-template.xlsx"

EXPECTED_SHEETS = [
    "Entities",
    "Relationships",
    "Names",
    "Relationship Guide",
    "Lookup Values",
    "Examples",
]
ENTRY_SHEETS: dict[str, set[str]] = {
    "Entities": {"breadcrumb"},
    "Relationships": set(),
    "Names": set(),
}


def normalized_header(value: object) -> str:
    """Normalize headers such as ``intake_key *`` to ``intake_key``."""
    return str(value or "").strip().removesuffix(" *").strip()


def clear_entry_sheet(sheet: Worksheet, preserved_columns: set[str]) -> None:
    """Clear entered records while preserving formatting and derived columns."""
    headers = {
        cell.column: normalized_header(cell.value) for cell in sheet[1] if cell.value
    }
    if not headers:
        raise ValueError(f"{sheet.title} does not contain a header row")
    unknown_preserved = preserved_columns - set(headers.values())
    if unknown_preserved:
        raise ValueError(
            f"{sheet.title} is missing preserved columns: "
            f"{', '.join(sorted(unknown_preserved))}"
        )

    for row_number in range(2, sheet.max_row + 1):
        for column_number, header in headers.items():
            if header in preserved_columns:
                continue
            cell = sheet.cell(row_number, column_number)
            cell.value = None
            cell.comment = None
            cell.hyperlink = None


def validate_breadcrumb_formulas(sheet: Worksheet) -> None:
    """Ensure the blank template retains its calculated breadcrumb column."""
    headers = {
        normalized_header(cell.value): cell.column for cell in sheet[1] if cell.value
    }
    breadcrumb_column = headers.get("breadcrumb")
    if breadcrumb_column is None:
        raise ValueError("Entities is missing the breadcrumb column")
    missing_formulas = [
        cell.coordinate
        for row in range(2, sheet.max_row + 1)
        if (cell := sheet.cell(row, breadcrumb_column)).data_type != "f"
    ]
    if missing_formulas:
        preview = ", ".join(missing_formulas[:5])
        raise ValueError(f"Entities has missing breadcrumb formulas: {preview}")


def create_template(source: Path, output: Path) -> None:
    """Copy the current workbook design and remove entered graph records."""
    if source.resolve() == output.resolve():
        raise ValueError("Template output must be different from the source workbook")

    workbook = load_workbook(source, read_only=False, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ValueError(
            "Unexpected workbook sheets. Expected "
            f"{EXPECTED_SHEETS!r}, found {workbook.sheetnames!r}"
        )

    for sheet_name, preserved_columns in ENTRY_SHEETS.items():
        clear_entry_sheet(workbook[sheet_name], preserved_columns)
    validate_breadcrumb_formulas(workbook["Entities"])

    workbook.properties.creator = "PHNER contributors"
    workbook.properties.lastModifiedBy = "PHNER contributors"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(f"{output.suffix}.tmp")
    try:
        workbook.save(temporary)
        temporary.replace(output)
    finally:
        temporary.unlink(missing_ok=True)


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    """Parse command-line options."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help=f"workbook to copy (default: {DEFAULT_SOURCE.relative_to(PROJECT_ROOT)})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"template destination (default: {DEFAULT_OUTPUT.relative_to(PROJECT_ROOT)})",
    )
    return parser.parse_args(argv)


def main(argv: Iterable[str] | None = None) -> int:
    """Create the template and report its destination."""
    args = parse_args(argv)
    source = args.source.resolve()
    output = args.output.resolve()
    if not source.is_file():
        raise SystemExit(f"Source workbook not found: {source}")
    lock = source.with_name(f"~${source.name}")
    if lock.exists():
        raise SystemExit(
            f"Excel appears to have the source workbook open ({lock.name}). "
            "Save and close it before regenerating the template."
        )

    create_template(source, output)
    print(f"Wrote blank intake template to {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
