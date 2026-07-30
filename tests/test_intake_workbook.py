from __future__ import annotations

import subprocess
import sys
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

DATA_SHEETS = [
    "Entities",
    "Relationships",
    "Names",
]


def assert_excel_compatible_entry_sheets(path: Path) -> None:
    """Reject empty Excel Table parts, which desktop Excel removes as unreadable."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    for sheet_name in DATA_SHEETS:
        sheet = workbook[sheet_name]
        assert not sheet.tables
        assert sheet.auto_filter.ref is not None
    assert len(workbook["Relationship Guide"].tables) == 1
    with ZipFile(path) as archive:
        table_parts = [name for name in archive.namelist() if name.startswith("xl/tables/")]
    assert table_parts == ["xl/tables/table1.xml"]


def test_committed_intake_template_matches_generator(
    repository_root: Path,
    tmp_path: Path,
) -> None:
    generated = tmp_path / "phner-intake-template.xlsx"
    subprocess.run(
        [
            sys.executable,
            "scripts/generate_intake_workbook.py",
            "--output",
            str(generated),
        ],
        check=True,
    )

    generated_workbook = load_workbook(generated)
    committed_workbook = load_workbook(
        repository_root / "templates" / "phner-intake-template.xlsx"
    )
    assert generated_workbook.sheetnames == [
        "Entities",
        "Relationships",
        "Names",
        "Relationship Guide",
        "Lookup Values",
        "Examples",
    ]
    assert committed_workbook.sheetnames == generated_workbook.sheetnames
    for sheet_name in generated_workbook.sheetnames:
        generated_sheet = generated_workbook[sheet_name]
        committed_sheet = committed_workbook[sheet_name]
        assert (generated_sheet.max_row, generated_sheet.max_column) == (
            committed_sheet.max_row,
            committed_sheet.max_column,
        )
        for row in generated_sheet.iter_rows():
            for cell in row:
                assert cell.value == committed_sheet[cell.coordinate].value

    entities = generated_workbook["Entities"]
    assert entities["A1"].value == "intake_key *"
    assert entities["C1"].value == "entity_type *"
    assert entities["I1"].value == "breadcrumb"
    assert all(entities.cell(row, 1).value is None for row in range(2, 1002))
    assert all(entities.cell(row, 9).data_type == "f" for row in range(2, 1002))
    assert all(
        generated_workbook["Relationships"].cell(row, 1).value is None
        for row in range(2, 1002)
    )
    assert all(
        generated_workbook["Names"].cell(row, 1).value is None
        for row in range(2, generated_workbook["Names"].max_row + 1)
    )
    assert "EntityTypes" in generated_workbook.defined_names
    assert "RelationshipTypes" in generated_workbook.defined_names
    assert "EntityKeys" in generated_workbook.defined_names
    assert len(entities.data_validations.dataValidation) == 2
    assert (
        len(generated_workbook["Relationships"].data_validations.dataValidation) == 2
    )
    assert generated_workbook.properties.creator == "PHNER contributors"
    assert generated_workbook.properties.lastModifiedBy == "PHNER contributors"
    assert_excel_compatible_entry_sheets(generated)


def test_committed_template_can_be_opened(repository_root: Path) -> None:
    path = repository_root / "templates" / "phner-intake-template.xlsx"
    workbook = load_workbook(path, read_only=False, data_only=False)
    assert workbook.properties.title == "PHNER graph intake workbook"
    assert workbook["Entities"]["I1"].value == "breadcrumb"
    assert_excel_compatible_entry_sheets(path)


def test_versioned_working_workbook_can_be_opened(repository_root: Path) -> None:
    path = repository_root / "intake" / "PHNER-US-FED-DHHS.xlsx"
    workbook = load_workbook(path, read_only=False, data_only=False)
    assert workbook.properties.title == "PHNER graph intake workbook"
    assert workbook.sheetnames == [
        "Entities",
        "Relationships",
        "Names",
        "Relationship Guide",
        "Lookup Values",
        "Examples",
    ]
    entities = workbook["Entities"]
    assert entities["I1"].value == "breadcrumb"
    assert entities["I2"].data_type == "f"
    assert "parent_intake_key" in entities["I1"].comment.text
    entity_count = sum(
        1 for row in entities.iter_rows(min_row=2, min_col=1, max_col=1) if row[0].value
    )
    assert entity_count > 0
